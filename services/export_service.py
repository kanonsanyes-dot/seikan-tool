from __future__ import annotations
from io import BytesIO
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file
from models import Order, Schedule
from services import cache_service

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
HEADER_FONT = Font(bold=True)

def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)

def orders_excel(query):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "受注一覧"
    ws.append(["受注ID","品名","工程品名","出荷先","出荷日","数量","ステータス","データ品質"])
    for o in query.all():
        ws.append([o.order_id,o.product_name,o.process_product_name,o.customer,o.ship_date,o.quantity,o.status,o.data_quality])
    _style_header(ws)
    buf=BytesIO(); wb.save(buf); buf.seek(0); cache_service.clear()
    return send_file(buf, as_attachment=True, download_name=f"受注一覧_{datetime.now():%Y%m%d}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def summary_excel(monthly, customer, product):
    wb=openpyxl.Workbook()
    for title, data, headers in [("月別集計", monthly, ["月","数量"]),("客先別集計", customer, ["客先","数量"]),("品名別集計", product, ["品名","数量"] )]:
        ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet(title)
        ws.title = title
        ws.append(headers)
        for k,v in data:
            ws.append([k,v])
        _style_header(ws)
    buf=BytesIO(); wb.save(buf); buf.seek(0); cache_service.clear()
    return send_file(buf, as_attachment=True, download_name=f"受注データ_集計_{datetime.now():%Y%m%d}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def scheduler_excel(schedules):
    wb=openpyxl.Workbook()
    ws1=wb.active; ws1.title="スケジュール一覧"
    ws1.append(["受注ID","品名","客先","工程","開始日","終了日","日数","ステータス","品質異常"])
    monthly=defaultdict(int); summary=defaultdict(lambda: {"total":0,"done":0})
    for s in schedules:
        o=s.order
        ws1.append([s.order_id,o.product_name,o.customer,s.process_name,s.start_date,s.end_date,s.required_days,s.status,"⚠" if s.has_quality_issue else ""])
        monthly[(s.process_name, s.start_date.strftime("%Y-%m"))] += o.quantity
        summary[s.order_id]["total"] += 1
        if s.status == "完了": summary[s.order_id]["done"] += 1
    _style_header(ws1)
    ws2=wb.create_sheet("工程別負荷集計"); ws2.append(["工程","月","仕掛数量"])
    for (proc, ym), qty in sorted(monthly.items()): ws2.append([proc, ym, qty])
    _style_header(ws2)
    ws3=wb.create_sheet("受注サマリ"); ws3.append(["受注ID","工程数","完了数","進捗率"])
    for oid, d in sorted(summary.items()): ws3.append([oid,d["total"],d["done"], f"{(d['done']/d['total']*100 if d['total'] else 0):.1f}%"])
    _style_header(ws3)
    buf=BytesIO(); wb.save(buf); buf.seek(0); cache_service.clear()
    return send_file(buf, as_attachment=True, download_name=f"schedule_{datetime.now():%Y%m%d}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def capacity_report_excel(summary, risks):
    """会議用キャパ確認レポート Excel版"""
    from openpyxl.styles import Border, Side
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # Sheet1: 工程別負荷率
    ws1 = wb.active; ws1.title = "工程別負荷率"
    ws1.append([f"{summary.year}年{summary.month}月 工程別負荷率"])
    ws1["A1"].font = Font(bold=True, size=13)
    ws1.append([])
    ws1.append(["工程名", "必要工数(h)", "稼働時間(h)", "負荷率(%)", "不足時間(h)", "+20h残業", "+40h残業", "判定"])
    for p in summary.process_loads:
        s20 = max(0, round(p.required_hours - p.available_hours - 20, 1))
        s40 = max(0, round(p.required_hours - p.available_hours - 40, 1))
        ws1.append([
            p.process_name,
            p.required_hours,
            p.available_hours,
            p.load_rate,
            p.required_overtime if p.required_overtime > 0 else "-",
            "✓" if s20 <= 0 else f"不足{s20}h",
            "✓" if s40 <= 0 else f"不足{s40}h",
            {"critical":"要調整","overtime":"残業要","caution":"注意"}.get(p.status, "余裕あり"),
        ])
    _style_header_from_row(ws1, 3)

    # Sheet2: 遅延リスク
    ws2 = wb.create_sheet("遅延リスク")
    ws2.append(["受注ID", "品名", "客先", "出荷日", "残日数", "未完了工程", "遅延工程", "リスクレベル"])
    for r in risks:
        ws2.append([
            r.order_id,
            r.product_name,
            r.customer,
            r.ship_date,
            r.days_until_ship,
            " / ".join(r.incomplete_processes),
            " / ".join(r.overdue_processes),
            {"overdue":"出荷超過","critical":"緊急(3日以内)","high":"高(7日以内)","medium":"中(工程遅延)"}.get(r.level, r.level),
        ])
    _style_header(ws2)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"キャパ確認レポート_{summary.year}{summary.month:02d}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _style_header_from_row(ws, row_num):
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)


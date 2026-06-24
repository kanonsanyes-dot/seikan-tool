from __future__ import annotations
from datetime import datetime, date
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from database import db
from models import MeasurementRecord

meas_bp = Blueprint("measurements", __name__, url_prefix="/measurements")

RESULT_LIST = ["OK", "NG", "要確認"]


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


@meas_bp.route("")
def list_measurements():
    q = request.args.get("q", "").strip()
    result = request.args.get("result", "").strip()
    date_from = _parse_date(request.args.get("date_from", ""))
    date_to = _parse_date(request.args.get("date_to", ""))
    query = MeasurementRecord.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(MeasurementRecord.product_name.like(like),
                   MeasurementRecord.lot_no.like(like),
                   MeasurementRecord.measurement_item.like(like))
        )
    if result:
        query = query.filter(MeasurementRecord.result == result)
    if date_from:
        query = query.filter(MeasurementRecord.measured_date >= date_from)
    if date_to:
        query = query.filter(MeasurementRecord.measured_date <= date_to)
    items = query.order_by(MeasurementRecord.measured_date.desc().nullslast(),
                           MeasurementRecord.measurement_id.desc()).all()
    ng_count = sum(1 for i in items if i.result == "NG")
    return render_template("measurements/list.html",
                           items=items, q=q, result=result,
                           date_from=request.args.get("date_from", ""),
                           date_to=request.args.get("date_to", ""),
                           ng_count=ng_count, result_list=RESULT_LIST)


@meas_bp.route("/new", methods=["GET", "POST"])
def new_measurement():
    work_order_id = request.args.get("work_order_id", type=int)
    process_id = request.args.get("process_id", type=int)
    if request.method == "POST":
        m = MeasurementRecord(
            work_order_id=request.form.get("work_order_id", type=int),
            process_id=request.form.get("process_id", type=int),
            product_name=request.form.get("product_name", "").strip(),
            lot_no=request.form.get("lot_no", "").strip(),
            measured_date=_parse_date(request.form.get("measured_date")),
            measurement_item=request.form.get("measurement_item", "").strip(),
            result=request.form.get("result", "").strip(),
            file_path=request.form.get("file_path", "").strip(),
            storage_note=request.form.get("storage_note", "").strip(),
            inspector=request.form.get("inspector", "").strip(),
            remarks=request.form.get("remarks", "").strip(),
        )
        if not m.product_name:
            flash("品名は必須です。", "danger")
            return render_template("measurements/form.html", m=None,
                                   work_order_id=work_order_id, process_id=process_id,
                                   result_list=RESULT_LIST)
        db.session.add(m)
        try:
            db.session.commit()
            flash("測定記録を登録しました。", "success")
            if m.work_order_id:
                return redirect(url_for("work_orders.detail", wo_id=m.work_order_id))
            return redirect(url_for("measurements.list_measurements"))
        except Exception as e:
            db.session.rollback()
            flash(f"登録エラー: {e}", "danger")
    return render_template("measurements/form.html", m=None,
                           work_order_id=work_order_id, process_id=process_id,
                           result_list=RESULT_LIST)


@meas_bp.route("/<int:meas_id>")
def detail(meas_id):
    m = MeasurementRecord.query.get_or_404(meas_id)
    return render_template("measurements/detail.html", m=m)


@meas_bp.route("/<int:meas_id>/edit", methods=["GET", "POST"])
def edit_measurement(meas_id):
    m = MeasurementRecord.query.get_or_404(meas_id)
    if request.method == "POST":
        m.product_name = request.form.get("product_name", "").strip()
        m.lot_no = request.form.get("lot_no", "").strip()
        m.measured_date = _parse_date(request.form.get("measured_date"))
        m.measurement_item = request.form.get("measurement_item", "").strip()
        m.result = request.form.get("result", "").strip()
        m.file_path = request.form.get("file_path", "").strip()
        m.storage_note = request.form.get("storage_note", "").strip()
        m.inspector = request.form.get("inspector", "").strip()
        m.remarks = request.form.get("remarks", "").strip()
        try:
            db.session.commit()
            flash("更新しました。", "success")
            return redirect(url_for("measurements.detail", meas_id=meas_id))
        except Exception as e:
            db.session.rollback()
            flash(f"更新エラー: {e}", "danger")
    return render_template("measurements/form.html", m=m,
                           work_order_id=m.work_order_id, process_id=m.process_id,
                           result_list=RESULT_LIST)


@meas_bp.route("/excel")
def excel_export():
    q = request.args.get("q", "").strip()
    result = request.args.get("result", "").strip()
    date_from = _parse_date(request.args.get("date_from", ""))
    date_to = _parse_date(request.args.get("date_to", ""))
    query = MeasurementRecord.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(MeasurementRecord.product_name.like(like),
                   MeasurementRecord.lot_no.like(like),
                   MeasurementRecord.measurement_item.like(like))
        )
    if result:
        query = query.filter(MeasurementRecord.result == result)
    if date_from:
        query = query.filter(MeasurementRecord.measured_date >= date_from)
    if date_to:
        query = query.filter(MeasurementRecord.measured_date <= date_to)
    items = query.order_by(MeasurementRecord.measured_date.desc().nullslast(),
                           MeasurementRecord.measurement_id.desc()).all()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "測定記録"
        hdr_fill = PatternFill("solid", fgColor="2563EB")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ["測定日", "品名", "ロット番号", "測定項目", "結果", "検査者", "備考", "保管メモ"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        for m in items:
            ws.append([
                str(m.measured_date) if m.measured_date else "",
                m.product_name,
                m.lot_no or "",
                m.measurement_item or "",
                m.result or "",
                m.inspector or "",
                m.remarks or "",
                m.storage_note or "",
            ])
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).border = border
        for col, width in zip("ABCDEFGH", [12, 24, 16, 22, 8, 14, 24, 24]):
            ws.column_dimensions[col].width = width
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"測定記録_{date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"Excel出力エラー: {e}", "danger")
        return redirect(url_for("measurements.list_measurements"))

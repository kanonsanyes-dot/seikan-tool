from __future__ import annotations
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from database import db, commit_or_rollback
from models import Customer, Product, ProductProcessStandard, ProcessCapacity, ProcessMaster, DEFAULT_PROCESSES, Order
from services.quality_check_service import recheck_all
from services import cache_service
from datetime import datetime

masters_bp=Blueprint("masters", __name__, url_prefix="/masters")

@masters_bp.route("/customers")
def customers():
    return render_template("masters/customers.html", customers=Customer.query.order_by(Customer.customer_name).all())

@masters_bp.route("/customers/add", methods=["POST"])
def add_customer():
    name=request.form.get("customer_name","").strip()
    if name and not Customer.query.filter_by(customer_name=name).first():
        db.session.add(Customer(customer_name=name, is_active=True)); commit_or_rollback(); cache_service.clear(); flash("出荷先を追加しました。", "success")
    return redirect(url_for("masters.customers"))

@masters_bp.route("/customers/<int:id>/edit", methods=["POST"])
def edit_customer(id):
    c=db.session.get(Customer,id)
    if c:
        c.customer_name=request.form.get("customer_name",c.customer_name); c.is_active=bool(request.form.get("is_active")); commit_or_rollback(); cache_service.clear(); recheck_all(Order.query.all()); commit_or_rollback()
        flash("出荷先を更新しました。", "success")
    return redirect(url_for("masters.customers"))

@masters_bp.route("/customers/<int:id>/delete", methods=["POST"])
def delete_customer(id):
    c=db.session.get(Customer,id)
    if c: db.session.delete(c); commit_or_rollback(); cache_service.clear(); flash("出荷先を削除しました。", "success")
    return redirect(url_for("masters.customers"))

@masters_bp.route("/products")
def products():
    return render_template("masters/products.html", products=Product.query.order_by(Product.product_name).all())

@masters_bp.route("/products/add", methods=["POST"])
def add_product():
    name=request.form.get("product_name","").strip()
    if name and not Product.query.filter_by(product_name=name).first():
        db.session.add(Product(product_name=name, process_product_name=request.form.get("process_product_name"), is_active=True)); commit_or_rollback(); cache_service.clear(); flash("品名を追加しました。", "success")
    return redirect(url_for("masters.products"))

@masters_bp.route("/products/<int:id>/edit", methods=["POST"])
def edit_product(id):
    p=db.session.get(Product,id)
    if p:
        p.product_name=request.form.get("product_name",p.product_name); p.process_product_name=request.form.get("process_product_name"); p.is_active=bool(request.form.get("is_active")); commit_or_rollback(); cache_service.clear(); recheck_all(Order.query.all()); commit_or_rollback(); flash("品名を更新しました。", "success")
    return redirect(url_for("masters.products"))

@masters_bp.route("/products/<int:id>/delete", methods=["POST"])
def delete_product(id):
    p=db.session.get(Product,id)
    if p: db.session.delete(p); commit_or_rollback(); cache_service.clear(); flash("品名を削除しました。", "success")
    return redirect(url_for("masters.products"))

@masters_bp.route("/process-standards")
def standards():
    product_filter = request.args.get("product", "")
    q = ProductProcessStandard.query.order_by(ProductProcessStandard.product_name, ProductProcessStandard.process_order)
    if product_filter:
        q = q.filter(ProductProcessStandard.product_name == product_filter)
    standards_list = q.all()
    product_names = [r[0] for r in db.session.query(ProductProcessStandard.product_name).distinct().order_by(ProductProcessStandard.product_name).all()]
    return render_template("masters/process_standards.html",
                           standards=standards_list,
                           product_names=product_names,
                           product_filter=product_filter)

@masters_bp.route("/process-standards/add", methods=["POST"])
def add_standard():
    product_name=request.form.get("product_name", "").strip()
    process_name=request.form.get("process_name", "").strip()
    if not product_name or not process_name:
        flash("品名と工程名は必須です。", "danger")
        return redirect(url_for("masters.standards"))
    if ProductProcessStandard.query.filter_by(product_name=product_name, process_name=process_name).first():
        flash("同じ品名・工程名の工程標準は既に登録されています。", "warning")
        return redirect(url_for("masters.standards"))
    try:
        s=ProductProcessStandard(product_name=product_name, process_product_name=request.form.get("process_product_name"), process_name=process_name, process_order=int(request.form.get("process_order",1)), standard_time_min=float(request.form.get("standard_time_min",0) or 0), daily_capacity=int(request.form.get("daily_capacity",0) or 0), lot_size=int(request.form.get("lot_size",0) or 0), is_active=True, remarks=request.form.get("remarks"))
        db.session.add(s); commit_or_rollback(); cache_service.clear(); flash("工程標準を追加しました。", "success")
    except Exception as e:
        db.session.rollback(); flash(f"工程標準の追加に失敗しました: {e}", "danger")
    return redirect(url_for("masters.standards"))

@masters_bp.route("/process-standards/<int:id>/delete", methods=["POST"])
def delete_standard(id):
    s=db.session.get(ProductProcessStandard,id)
    if s: db.session.delete(s); commit_or_rollback(); cache_service.clear(); flash("工程標準を削除しました。", "success")
    return redirect(url_for("masters.standards"))

@masters_bp.route("/process-capacity")
def capacity():
    proc_filter = request.args.get("process", "")
    date_from   = request.args.get("date_from", "")
    date_to     = request.args.get("date_to", "")
    q = ProcessCapacity.query.order_by(ProcessCapacity.work_date.desc(), ProcessCapacity.process_name)
    if proc_filter:
        q = q.filter(ProcessCapacity.process_name == proc_filter)
    if date_from:
        try: q = q.filter(ProcessCapacity.work_date >= datetime.fromisoformat(date_from).date())
        except ValueError: pass
    if date_to:
        try: q = q.filter(ProcessCapacity.work_date <= datetime.fromisoformat(date_to).date())
        except ValueError: pass
    capacities = q.limit(500).all()
    process_names = [r[0] for r in db.session.query(ProcessCapacity.process_name).distinct().order_by(ProcessCapacity.process_name).all()]
    return render_template("masters/process_capacity.html",
                           capacities=capacities,
                           process_names=process_names,
                           proc_filter=proc_filter,
                           date_from=date_from,
                           date_to=date_to)

@masters_bp.route("/process-capacity/add", methods=["POST"])
def add_capacity():
    process_name=request.form.get("process_name", "").strip()
    work_date_raw=request.form.get("work_date", "")
    if not process_name or not work_date_raw:
        flash("工程名と稼働日は必須です。", "danger")
        return redirect(url_for("masters.capacity"))
    try:
        work_date=datetime.fromisoformat(work_date_raw).date()
        if ProcessCapacity.query.filter_by(process_name=process_name, work_date=work_date).first():
            flash("同じ工程名・稼働日の工程キャパは既に登録されています。", "warning")
            return redirect(url_for("masters.capacity"))
        c=ProcessCapacity(process_name=process_name, work_date=work_date, available_hours=float(request.form.get("available_hours",8) or 8), workers=int(request.form.get("workers",1) or 1), overtime_hours=float(request.form.get("overtime_hours",0) or 0), capacity_quantity=int(request.form.get("capacity_quantity",0) or 0), remarks=request.form.get("remarks"))
        db.session.add(c); commit_or_rollback(); cache_service.clear(); flash("工程キャパを追加しました。", "success")
    except Exception as e:
        db.session.rollback(); flash(f"工程キャパの追加に失敗しました: {e}", "danger")
    return redirect(url_for("masters.capacity"))

@masters_bp.route("/process-capacity/<int:id>/delete", methods=["POST"])
def delete_capacity(id):
    c=db.session.get(ProcessCapacity,id)
    if c: db.session.delete(c); commit_or_rollback(); cache_service.clear(); flash("工程キャパを削除しました。", "success")
    return redirect(url_for("masters.capacity"))


# ─── 工程マスタ ──────────────────────────────────────────────────
@masters_bp.route("/process-masters")
def process_masters():
    masters = ProcessMaster.query.order_by(ProcessMaster.display_order).all()
    return render_template("masters/process_masters.html",
                           masters=masters, default_processes=DEFAULT_PROCESSES)

@masters_bp.route("/process-masters/add", methods=["POST"])
def add_process_master():
    name = request.form.get("process_name", "").strip()
    if not name:
        flash("工程名は必須です。", "danger")
        return redirect(url_for("masters.process_masters"))
    if ProcessMaster.query.filter_by(process_name=name).first():
        flash("同名の工程が既に存在します。", "warning")
        return redirect(url_for("masters.process_masters"))
    last = ProcessMaster.query.order_by(ProcessMaster.display_order.desc()).first()
    order = (last.display_order + 1) if last else 1
    pm = ProcessMaster(
        process_name=name, display_order=order,
        hours_per_day=float(request.form.get("hours_per_day", 8) or 8),
        overtime_hours=float(request.form.get("overtime_hours", 0) or 0),
        pace_per_hour=int(request.form.get("pace_per_hour", 0) or 0),
        is_active=True,
        remarks=request.form.get("remarks", ""),
    )
    db.session.add(pm)
    commit_or_rollback()
    flash(f"工程「{name}」を追加しました。", "success")
    return redirect(url_for("masters.process_masters"))

@masters_bp.route("/process-masters/<int:pid>/edit", methods=["POST"])
def edit_process_master(pid):
    pm = db.session.get(ProcessMaster, pid)
    if not pm:
        flash("工程が見つかりません。", "danger")
        return redirect(url_for("masters.process_masters"))
    pm.process_name = request.form.get("process_name", pm.process_name).strip()
    pm.hours_per_day = float(request.form.get("hours_per_day", pm.hours_per_day) or pm.hours_per_day)
    pm.overtime_hours = float(request.form.get("overtime_hours", 0) or 0)
    pm.pace_per_hour = int(request.form.get("pace_per_hour", 0) or 0)
    pm.is_active = bool(request.form.get("is_active"))
    pm.remarks = request.form.get("remarks", "")
    commit_or_rollback()
    flash("工程を更新しました。", "success")
    return redirect(url_for("masters.process_masters"))

@masters_bp.route("/process-masters/<int:pid>/delete", methods=["POST"])
def delete_process_master(pid):
    pm = db.session.get(ProcessMaster, pid)
    if pm:
        db.session.delete(pm); commit_or_rollback()
        flash("工程を削除しました。", "success")
    return redirect(url_for("masters.process_masters"))

@masters_bp.route("/process-masters/import-excel", methods=["POST"])
def import_capacity_excel():
    from pathlib import Path
    f = request.files.get("file")
    if not f or not f.filename:
        flash("ファイルを選択してください。", "danger")
        return redirect(url_for("masters.process_masters"))
    fpath = Path("/tmp") / "capacity_import.xlsx"
    f.save(str(fpath))
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(fpath), data_only=True)
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            name = str(row[0]).strip()
            hours = float(row[1] or 8)
            overtime = float(row[2] or 0)
            pace = int(row[3] or 0) if len(row) > 3 else 0
            pm = ProcessMaster.query.filter_by(process_name=name).first()
            if pm:
                pm.hours_per_day = hours; pm.overtime_hours = overtime; pm.pace_per_hour = pace
            else:
                last = ProcessMaster.query.order_by(ProcessMaster.display_order.desc()).first()
                ord_ = (last.display_order + 1) if last else 1
                db.session.add(ProcessMaster(
                    process_name=name, display_order=ord_,
                    hours_per_day=hours, overtime_hours=overtime, pace_per_hour=pace, is_active=True
                ))
            imported += 1
        commit_or_rollback()
        flash(f"{imported} 件の工程キャパシティを取込しました。", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"インポートエラー: {e}", "danger")
    return redirect(url_for("masters.process_masters"))


# ─── 品名工程テンプレート ──────────────────────────────────────────
@masters_bp.route("/product-templates")
def product_templates():
    p_masters = ProcessMaster.query.filter_by(is_active=True).order_by(ProcessMaster.display_order).all()
    products = db.session.query(ProductProcessStandard.product_name).distinct().all()
    product_names = [p[0] for p in products]
    from models import Product
    all_products = Product.query.filter_by(is_active=True).order_by(Product.product_name).all()
    for ap in all_products:
        if ap.product_name not in product_names:
            product_names.append(ap.product_name)
    return render_template("masters/product_templates.html",
                           process_masters=p_masters,
                           product_names=sorted(product_names))

@masters_bp.route("/product-templates/<path:product_name>")
def product_template_detail(product_name):
    p_masters = ProcessMaster.query.filter_by(is_active=True).order_by(ProcessMaster.display_order).all()
    standards = {s.process_name: s for s in
                 ProductProcessStandard.query.filter_by(product_name=product_name).all()}
    return render_template("masters/product_template_edit.html",
                           product_name=product_name,
                           process_masters=p_masters,
                           standards=standards)

@masters_bp.route("/product-templates/<path:product_name>/save", methods=["POST"])
def save_product_template(product_name):
    p_masters = ProcessMaster.query.filter_by(is_active=True).order_by(ProcessMaster.display_order).all()
    selected = set(request.form.getlist("processes"))
    for i, pm in enumerate(p_masters, start=1):
        existing = ProductProcessStandard.query.filter_by(
            product_name=product_name, process_name=pm.process_name).first()
        if pm.process_name in selected:
            hours = float(request.form.get(f"hours_{pm.process_id}", pm.hours_per_day) or pm.hours_per_day)
            pace = int(request.form.get(f"pace_{pm.process_id}", pm.pace_per_hour) or 0)
            daily = int(hours * pace) if pace else 0
            if existing:
                existing.process_order = i; existing.hours_per_run = hours
                existing.pace_per_hour = pace; existing.daily_capacity = daily; existing.is_active = True
            else:
                db.session.add(ProductProcessStandard(
                    product_name=product_name, process_name=pm.process_name,
                    process_order=i, hours_per_run=hours, pace_per_hour=pace,
                    daily_capacity=daily, is_active=True,
                ))
        else:
            if existing:
                db.session.delete(existing)
    commit_or_rollback()
    flash(f"「{product_name}」の工程テンプレートを保存しました。", "success")
    return redirect(url_for("masters.product_template_detail", product_name=product_name))

@masters_bp.route("/process-masters/capacity-template")
def download_capacity_template():
    from io import BytesIO
    from flask import send_file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "工程キャパ"
        hdr_fill = PatternFill("solid", fgColor="2563EB")
        hdr_font = Font(bold=True, color="FFFFFF")
        headers = ["工程名", "定時h/日", "残業h/日", "pcs/h（ペース）"]
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        from models import ProcessMaster
        for pm in ProcessMaster.query.order_by(ProcessMaster.display_order).all():
            ws.append([pm.process_name, pm.hours_per_day, pm.overtime_hours, pm.pace_per_hour])
        for col, width in zip("ABCD", [20, 12, 12, 16]):
            ws.column_dimensions[col].width = width
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True, download_name="工程キャパシティ_テンプレート.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"テンプレート生成エラー: {e}", "danger")
        return redirect(url_for("masters.process_masters"))


# ─── 共通ヘルパー ─────────────────────────────────────────────────

def _make_wb(title, headers, rows, col_widths=None):
    """ヘッダ付きスタイル済みワークブックを返す"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = title
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(1, i)
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    if col_widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return wb


def _wb_to_response(wb, filename):
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _parse_upload(f):
    """アップロードされたファイル（xlsx / csv）を [[セル, ...], ...] として返す (ヘッダ行除く)"""
    import os, csv
    fname = (f.filename or "").lower()
    raw = f.read()
    if fname.endswith(".csv"):
        import io
        for enc in ("utf-8-sig", "cp932", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        return rows[1:] if rows else []
    else:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return [list(r) for r in rows]


def _str(v, default=""):
    return str(v).strip() if v is not None else default


def _float(v, default=0.0):
    try: return float(v)
    except (TypeError, ValueError): return default


def _int(v, default=0):
    try: return int(float(v))
    except (TypeError, ValueError): return default


# ─── 出荷先マスタ Import / Export ─────────────────────────────────

# ─── 一括削除 ────────────────────────────────────────────────────

@masters_bp.route("/customers/bulk-delete", methods=["POST"])
def bulk_delete_customers():
    ids = request.form.getlist("ids", type=int)
    if ids:
        Customer.query.filter(Customer.customer_id.in_(ids)).delete(synchronize_session=False)
        commit_or_rollback(); cache_service.clear()
        flash(f"出荷先 {len(ids)}件を削除しました。", "success")
    return redirect(url_for("masters.customers"))

@masters_bp.route("/products/bulk-delete", methods=["POST"])
def bulk_delete_products():
    ids = request.form.getlist("ids", type=int)
    if ids:
        Product.query.filter(Product.product_id.in_(ids)).delete(synchronize_session=False)
        commit_or_rollback(); cache_service.clear()
        flash(f"品名 {len(ids)}件を削除しました。", "success")
    return redirect(url_for("masters.products"))

@masters_bp.route("/process-standards/bulk-delete", methods=["POST"])
def bulk_delete_standards():
    ids = request.form.getlist("ids", type=int)
    if ids:
        ProductProcessStandard.query.filter(ProductProcessStandard.standard_id.in_(ids)).delete(synchronize_session=False)
        commit_or_rollback(); cache_service.clear()
        flash(f"工程標準 {len(ids)}件を削除しました。", "success")
    return redirect(url_for("masters.standards"))

@masters_bp.route("/process-capacity/bulk-delete", methods=["POST"])
def bulk_delete_capacity():
    ids = request.form.getlist("ids", type=int)
    if ids:
        ProcessCapacity.query.filter(ProcessCapacity.capacity_id.in_(ids)).delete(synchronize_session=False)
        commit_or_rollback(); cache_service.clear()
        flash(f"工程キャパ {len(ids)}件を削除しました。", "success")
    return redirect(url_for("masters.capacity"))

@masters_bp.route("/process-masters/bulk-delete", methods=["POST"])
def bulk_delete_process_masters():
    ids = request.form.getlist("ids", type=int)
    if ids:
        ProcessMaster.query.filter(ProcessMaster.process_id.in_(ids)).delete(synchronize_session=False)
        commit_or_rollback(); cache_service.clear()
        flash(f"工程マスタ {len(ids)}件を削除しました。", "success")
    return redirect(url_for("masters.process_masters"))


# ─── 出荷先マスタ Import / Export ─────────────────────────────────

@masters_bp.route("/customers/template")
def customers_template():
    rows = [[c.customer_name, "1" if c.is_active else "0"]
            for c in Customer.query.order_by(Customer.customer_name).all()]
    wb = _make_wb("出荷先マスタ",
                  ["出荷先名", "有効(1=有効/0=無効)"],
                  rows, [30, 18])
    return _wb_to_response(wb, "出荷先マスタ_テンプレート.xlsx")


@masters_bp.route("/customers/import", methods=["POST"])
def import_customers():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("ファイルを選択してください。", "danger")
        return redirect(url_for("masters.customers"))
    try:
        added = updated = skipped = 0
        for row in _parse_upload(f):
            name = _str(row[0] if row else "")
            if not name:
                continue
            active = _int(row[1] if len(row) > 1 else 1, 1) != 0
            existing = Customer.query.filter_by(customer_name=name).first()
            if existing:
                existing.is_active = active; updated += 1
            else:
                db.session.add(Customer(customer_name=name, is_active=active)); added += 1
        commit_or_rollback(); cache_service.clear()
        flash(f"出荷先マスタ取込: 追加 {added}件 / 更新 {updated}件", "success")
    except Exception as e:
        db.session.rollback(); flash(f"取込エラー: {e}", "danger")
    return redirect(url_for("masters.customers"))


# ─── 品名マスタ Import / Export ───────────────────────────────────

@masters_bp.route("/products/template")
def products_template():
    rows = [[p.product_name, p.process_product_name or "", "1" if p.is_active else "0"]
            for p in Product.query.order_by(Product.product_name).all()]
    wb = _make_wb("品名マスタ",
                  ["品名", "工程品名", "有効(1=有効/0=無効)"],
                  rows, [30, 30, 18])
    return _wb_to_response(wb, "品名マスタ_テンプレート.xlsx")


@masters_bp.route("/products/import", methods=["POST"])
def import_products():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("ファイルを選択してください。", "danger")
        return redirect(url_for("masters.products"))
    try:
        added = updated = 0
        for row in _parse_upload(f):
            name = _str(row[0] if row else "")
            if not name:
                continue
            proc_name = _str(row[1] if len(row) > 1 else "")
            active = _int(row[2] if len(row) > 2 else 1, 1) != 0
            existing = Product.query.filter_by(product_name=name).first()
            if existing:
                existing.process_product_name = proc_name or existing.process_product_name
                existing.is_active = active; updated += 1
            else:
                db.session.add(Product(product_name=name, process_product_name=proc_name or None, is_active=active))
                added += 1
        commit_or_rollback(); cache_service.clear()
        flash(f"品名マスタ取込: 追加 {added}件 / 更新 {updated}件", "success")
    except Exception as e:
        db.session.rollback(); flash(f"取込エラー: {e}", "danger")
    return redirect(url_for("masters.products"))


# ─── 工程標準 Import / Export ─────────────────────────────────────

@masters_bp.route("/process-standards/template")
def standards_template():
    rows = [[s.product_name, s.process_product_name or "", s.process_name,
             s.process_order, s.standard_time_min or 0, s.daily_capacity or 0,
             s.lot_size or 0, s.remarks or ""]
            for s in ProductProcessStandard.query.order_by(
                ProductProcessStandard.product_name,
                ProductProcessStandard.process_order).all()]
    wb = _make_wb("工程標準",
                  ["品名", "工程品名", "工程名", "工程順", "標準時間(分/個)", "日産数", "ロットサイズ", "備考"],
                  rows, [30, 30, 18, 8, 16, 10, 12, 20])
    return _wb_to_response(wb, "工程標準_テンプレート.xlsx")


@masters_bp.route("/process-standards/import", methods=["POST"])
def import_standards():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("ファイルを選択してください。", "danger")
        return redirect(url_for("masters.standards"))
    try:
        added = updated = 0
        for row in _parse_upload(f):
            product_name = _str(row[0] if row else "")
            process_name = _str(row[2] if len(row) > 2 else "")
            if not product_name or not process_name:
                continue
            existing = ProductProcessStandard.query.filter_by(
                product_name=product_name, process_name=process_name).first()
            if existing:
                existing.process_product_name = _str(row[1] if len(row) > 1 else "") or existing.process_product_name
                existing.process_order = _int(row[3] if len(row) > 3 else existing.process_order, existing.process_order)
                existing.standard_time_min = _float(row[4] if len(row) > 4 else 0)
                existing.daily_capacity = _int(row[5] if len(row) > 5 else 0)
                existing.lot_size = _int(row[6] if len(row) > 6 else 0)
                existing.remarks = _str(row[7] if len(row) > 7 else "")
                updated += 1
            else:
                db.session.add(ProductProcessStandard(
                    product_name=product_name,
                    process_product_name=_str(row[1] if len(row) > 1 else "") or None,
                    process_name=process_name,
                    process_order=_int(row[3] if len(row) > 3 else 1, 1),
                    standard_time_min=_float(row[4] if len(row) > 4 else 0),
                    daily_capacity=_int(row[5] if len(row) > 5 else 0),
                    lot_size=_int(row[6] if len(row) > 6 else 0),
                    remarks=_str(row[7] if len(row) > 7 else ""),
                    is_active=True,
                ))
                added += 1
        commit_or_rollback(); cache_service.clear()
        flash(f"工程標準取込: 追加 {added}件 / 更新 {updated}件", "success")
    except Exception as e:
        db.session.rollback(); flash(f"取込エラー: {e}", "danger")
    return redirect(url_for("masters.standards"))


# ─── 工程キャパ Import / Export ───────────────────────────────────

@masters_bp.route("/process-capacity/template")
def capacity_template():
    rows = [[c.process_name, c.work_date.isoformat() if c.work_date else "",
             c.available_hours or 8, c.overtime_hours or 0,
             c.workers or 1, c.capacity_quantity or 0, c.remarks or ""]
            for c in ProcessCapacity.query.order_by(
                ProcessCapacity.work_date.desc(),
                ProcessCapacity.process_name).limit(500).all()]
    wb = _make_wb("工程キャパ",
                  ["工程名", "稼働日(YYYY-MM-DD)", "定時h", "残業h", "人員", "処理数量", "備考"],
                  rows, [20, 20, 10, 10, 8, 12, 20])
    return _wb_to_response(wb, "工程キャパ_テンプレート.xlsx")


@masters_bp.route("/process-capacity/import", methods=["POST"])
def import_capacity():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("ファイルを選択してください。", "danger")
        return redirect(url_for("masters.capacity"))
    try:
        added = updated = skipped = 0
        for row in _parse_upload(f):
            process_name = _str(row[0] if row else "")
            date_raw = _str(row[1] if len(row) > 1 else "")
            if not process_name or not date_raw:
                skipped += 1; continue
            try:
                work_date = datetime.fromisoformat(date_raw).date()
            except ValueError:
                skipped += 1; continue
            existing = ProcessCapacity.query.filter_by(
                process_name=process_name, work_date=work_date).first()
            if existing:
                existing.available_hours = _float(row[2] if len(row) > 2 else 8, 8)
                existing.overtime_hours  = _float(row[3] if len(row) > 3 else 0)
                existing.workers         = _int(row[4] if len(row) > 4 else 1, 1)
                existing.capacity_quantity = _int(row[5] if len(row) > 5 else 0)
                existing.remarks         = _str(row[6] if len(row) > 6 else "")
                updated += 1
            else:
                db.session.add(ProcessCapacity(
                    process_name=process_name, work_date=work_date,
                    available_hours=_float(row[2] if len(row) > 2 else 8, 8),
                    overtime_hours=_float(row[3] if len(row) > 3 else 0),
                    workers=_int(row[4] if len(row) > 4 else 1, 1),
                    capacity_quantity=_int(row[5] if len(row) > 5 else 0),
                    remarks=_str(row[6] if len(row) > 6 else ""),
                ))
                added += 1
        commit_or_rollback(); cache_service.clear()
        msg = f"工程キャパ取込: 追加 {added}件 / 更新 {updated}件"
        if skipped: msg += f" / スキップ {skipped}件"
        flash(msg, "success")
    except Exception as e:
        db.session.rollback(); flash(f"取込エラー: {e}", "danger")
    return redirect(url_for("masters.capacity"))


# ─── 工程マスタ Import / Export ───────────────────────────────────

@masters_bp.route("/process-masters/template")
def process_masters_template():
    rows = [[pm.process_name, pm.hours_per_day, pm.overtime_hours,
             pm.pace_per_hour, pm.display_order, "1" if pm.is_active else "0", pm.remarks or ""]
            for pm in ProcessMaster.query.order_by(ProcessMaster.display_order).all()]
    wb = _make_wb("工程マスタ",
                  ["工程名", "定時h/日", "残業h/日", "pcs/h(ペース)", "表示順", "有効(1/0)", "備考"],
                  rows, [20, 10, 10, 14, 8, 10, 20])
    return _wb_to_response(wb, "工程マスタ_テンプレート.xlsx")


@masters_bp.route("/process-masters/import", methods=["POST"])
def import_process_masters():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("ファイルを選択してください。", "danger")
        return redirect(url_for("masters.process_masters"))
    try:
        added = updated = 0
        for row in _parse_upload(f):
            name = _str(row[0] if row else "")
            if not name:
                continue
            existing = ProcessMaster.query.filter_by(process_name=name).first()
            if existing:
                existing.hours_per_day   = _float(row[1] if len(row) > 1 else existing.hours_per_day, existing.hours_per_day)
                existing.overtime_hours  = _float(row[2] if len(row) > 2 else 0)
                existing.pace_per_hour   = _int(row[3] if len(row) > 3 else 0)
                if len(row) > 4 and row[4] is not None:
                    existing.display_order = _int(row[4], existing.display_order)
                existing.is_active       = _int(row[5] if len(row) > 5 else 1, 1) != 0
                existing.remarks         = _str(row[6] if len(row) > 6 else "")
                updated += 1
            else:
                last = ProcessMaster.query.order_by(ProcessMaster.display_order.desc()).first()
                ord_ = _int(row[4] if len(row) > 4 else None, (last.display_order + 1) if last else 1)
                db.session.add(ProcessMaster(
                    process_name=name,
                    hours_per_day=_float(row[1] if len(row) > 1 else 8, 8),
                    overtime_hours=_float(row[2] if len(row) > 2 else 0),
                    pace_per_hour=_int(row[3] if len(row) > 3 else 0),
                    display_order=ord_,
                    is_active=_int(row[5] if len(row) > 5 else 1, 1) != 0,
                    remarks=_str(row[6] if len(row) > 6 else ""),
                ))
                added += 1
        commit_or_rollback(); cache_service.clear()
        flash(f"工程マスタ取込: 追加 {added}件 / 更新 {updated}件", "success")
    except Exception as e:
        db.session.rollback(); flash(f"取込エラー: {e}", "danger")
    return redirect(url_for("masters.process_masters"))

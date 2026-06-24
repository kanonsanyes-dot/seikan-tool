from __future__ import annotations
from datetime import datetime, date
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from database import db
from models import WorkOrder, WorkOrderProcess, Order
from services.work_order_service import (
    create_work_order, update_work_order, generate_from_order,
    update_process_actual, get_work_order_list,
)

wo_bp = Blueprint("work_orders", __name__, url_prefix="/work-orders")

STATUS_LIST = ["未着手", "進行中", "完了", "保留", "停止", "取消"]
PRIORITY_LIST = ["緊急", "高", "通常", "低"]


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


@wo_bp.route("")
def list_work_orders():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    customer = request.args.get("customer", "").strip()
    date_from = _parse_date(request.args.get("date_from", ""))
    date_to = _parse_date(request.args.get("date_to", ""))
    items = get_work_order_list(q=q, status=status, customer=customer,
                                date_from=date_from, date_to=date_to)
    today = date.today()
    return render_template("work_orders/list.html",
                           items=items, today=today,
                           q=q, status=status, customer=customer,
                           date_from=request.args.get("date_from", ""),
                           date_to=request.args.get("date_to", ""),
                           status_list=STATUS_LIST)


@wo_bp.route("/new", methods=["GET", "POST"])
def new_work_order():
    if request.method == "POST":
        data = {
            "product_name": request.form.get("product_name", "").strip(),
            "process_product_name": request.form.get("process_product_name", "").strip(),
            "customer": request.form.get("customer", "").strip(),
            "lot_no": request.form.get("lot_no", "").strip(),
            "quantity": request.form.get("quantity", "0"),
            "ship_date": _parse_date(request.form.get("ship_date", "")),
            "status": request.form.get("status", "未着手"),
            "priority": request.form.get("priority", "通常"),
            "remarks": request.form.get("remarks", "").strip(),
        }
        if not data["product_name"]:
            flash("品名は必須です。", "danger")
            return render_template("work_orders/form.html", wo=None, data=data,
                                   status_list=STATUS_LIST, priority_list=PRIORITY_LIST)
        try:
            wo = create_work_order(data)
            flash(f"工程管理票 {wo.work_order_no} を作成しました。", "success")
            return redirect(url_for("work_orders.detail", wo_id=wo.work_order_id))
        except Exception as e:
            db.session.rollback()
            flash(f"登録エラー: {e}", "danger")
    return render_template("work_orders/form.html", wo=None, data={},
                           status_list=STATUS_LIST, priority_list=PRIORITY_LIST)


@wo_bp.route("/<int:wo_id>")
def detail(wo_id):
    wo = WorkOrder.query.get_or_404(wo_id)
    today = date.today()
    return render_template("work_orders/detail.html", wo=wo, today=today)


@wo_bp.route("/<int:wo_id>/edit", methods=["GET", "POST"])
def edit_work_order(wo_id):
    wo = WorkOrder.query.get_or_404(wo_id)
    if request.method == "POST":
        data = {
            "product_name": request.form.get("product_name", "").strip(),
            "process_product_name": request.form.get("process_product_name", "").strip(),
            "customer": request.form.get("customer", "").strip(),
            "lot_no": request.form.get("lot_no", "").strip(),
            "quantity": request.form.get("quantity", str(wo.quantity)),
            "ship_date": _parse_date(request.form.get("ship_date", "")),
            "status": request.form.get("status", wo.status),
            "priority": request.form.get("priority", wo.priority),
            "remarks": request.form.get("remarks", "").strip(),
        }
        try:
            update_work_order(wo, data)
            flash("更新しました。", "success")
            return redirect(url_for("work_orders.detail", wo_id=wo_id))
        except Exception as e:
            db.session.rollback()
            flash(f"更新エラー: {e}", "danger")
    return render_template("work_orders/form.html", wo=wo, data={},
                           status_list=STATUS_LIST, priority_list=PRIORITY_LIST)


@wo_bp.route("/<int:wo_id>/actuals", methods=["GET", "POST"])
def actuals(wo_id):
    wo = WorkOrder.query.get_or_404(wo_id)
    if request.method == "POST":
        for proc in wo.processes:
            key = str(proc.process_id)
            data = {
                "actual_start_date": request.form.get(f"start_{key}", ""),
                "actual_end_date": request.form.get(f"end_{key}", ""),
                "input_quantity": request.form.get(f"input_{key}", ""),
                "good_quantity": request.form.get(f"good_{key}", ""),
                "defect_quantity": request.form.get(f"defect_{key}", ""),
                "status": request.form.get(f"status_{key}", proc.status),
                "operator": request.form.get(f"operator_{key}", ""),
                "equipment": request.form.get(f"equipment_{key}", ""),
                "remarks": request.form.get(f"remarks_{key}", ""),
            }
            update_process_actual(proc, data)
        flash("実績を保存しました。", "success")
        return redirect(url_for("work_orders.actuals", wo_id=wo_id))
    return render_template("work_orders/actuals.html", wo=wo,
                           status_list=STATUS_LIST)


@wo_bp.route("/<int:wo_id>/print")
def print_view(wo_id):
    wo = WorkOrder.query.get_or_404(wo_id)
    today = date.today()
    return render_template("work_orders/print.html", wo=wo, today=today)


@wo_bp.route("/<int:wo_id>/excel")
def excel_export(wo_id):
    wo = WorkOrder.query.get_or_404(wo_id)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "工程管理票"

        hdr_fill = PatternFill("solid", fgColor="2563EB")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        info = [
            ["管理番号", wo.work_order_no, "状態", wo.status],
            ["品名", wo.product_name, "優先度", wo.priority],
            ["顧客", wo.customer or "", "ロット", wo.lot_no or ""],
            ["数量", wo.quantity, "納期", str(wo.ship_date) if wo.ship_date else ""],
        ]
        for row in info:
            ws.append(row)

        ws.append([])
        headers = ["工程名", "順序", "計画開始", "計画終了", "実績開始", "実績終了",
                   "投入数", "良品数", "不良数", "歩留(%)", "状態", "作業者", "備考"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border

        for p in wo.processes:
            yield_rate = round(p.good_quantity / p.input_quantity * 100, 1) if p.input_quantity else ""
            ws.append([
                p.process_name, p.process_order,
                str(p.planned_start_date or ""), str(p.planned_end_date or ""),
                str(p.actual_start_date or ""), str(p.actual_end_date or ""),
                p.input_quantity, p.good_quantity, p.defect_quantity,
                yield_rate, p.status, p.operator or "", p.remarks or "",
            ])
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).border = border

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 6

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"工程管理票_{wo.work_order_no}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"Excel出力エラー: {e}", "danger")
        return redirect(url_for("work_orders.detail", wo_id=wo_id))


@wo_bp.route("/bulk-status", methods=["POST"])
def bulk_status():
    ids = request.form.getlist("wo_ids")
    new_status = request.form.get("bulk_status", "").strip()
    if not ids or not new_status:
        flash("対象と変更先ステータスを選択してください。", "warning")
        return redirect(url_for("work_orders.list_work_orders"))
    updated = 0
    for wo_id in ids:
        wo = WorkOrder.query.get(int(wo_id))
        if wo:
            wo.status = new_status
            updated += 1
    try:
        db.session.commit()
        flash(f"{updated} 件のステータスを「{new_status}」に変更しました。", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"一括変更エラー: {e}", "danger")
    return redirect(url_for("work_orders.list_work_orders"))


@wo_bp.route("/generate-from-order/<int:order_id>", methods=["POST"])
def generate_from_order_view(order_id):
    order = Order.query.get_or_404(order_id)
    wo = generate_from_order(order)
    if wo is None:
        flash("この受注の工程管理票は既に生成されています。", "warning")
        existing = WorkOrder.query.filter_by(order_id=order_id).first()
        if existing:
            return redirect(url_for("work_orders.detail", wo_id=existing.work_order_id))
    else:
        flash(f"工程管理票 {wo.work_order_no} を生成しました。", "success")
        return redirect(url_for("work_orders.detail", wo_id=wo.work_order_id))
    return redirect(url_for("orders.list_orders"))
